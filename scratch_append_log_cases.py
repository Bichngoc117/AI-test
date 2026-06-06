import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# LOG TEST CASES từ BD Section 7. Log
# ============================================================
# Log format:
# User :logged_in_user_id <action>. inspection_id=:id :access_url :date_time :ip_address :user_agent

LOG_EXPECTED_FIELDS = "inspection_id, access_url, date_time, ip_address, user_agent"

log_cases = [
    # L01 - Vào Master Room
    {
        "TC No.": "MINSP_005_LOG_001",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Access Master Room",
        "Check content": "LOG - Ghi log khi vào màn hình Master Room",
        "Pre-condition / Test Data": "User đã login với role Engineer.\nInspection tồn tại và user được assign.",
        "Steps": "1. Vào màn hình MOB-010 Inspection Execute (Master Room)\n2. Kiểm tra log hệ thống",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} access inspection execute screen - master room. inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Đầy đủ 5 trường: inspection_id, access_url, date_time, ip_address, user_agent\n- logged_in_user_id khớp user đang login\n- inspection_id khớp inspection đang thực hiện",
        "Priority": "High",
    },

    # L02 - Download site location thành công
    {
        "TC No.": "MINSP_005_LOG_002",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Download - Site Location",
        "Check content": "LOG - Ghi log khi Download dữ liệu site location THÀNH CÔNG",
        "Pre-condition / Test Data": "Đang ở tab một phòng (site_location).\nĐang Online.\nCó data trên server để download.",
        "Steps": "1. Click nút ダウンロード (Download) tại tab site_location\n2. Chọn YES tại popup xác nhận\n3. Download thành công\n4. Kiểm tra log hệ thống",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} download data inspection result successful for site_location_id:{site_location_id} inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Chứa đúng site_location_id của phòng đang download\n- Đầy đủ các trường: inspection_id, access_url, date_time, ip_address, user_agent",
        "Priority": "High",
    },

    # L03 - Download site location thất bại
    {
        "TC No.": "MINSP_005_LOG_003",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Download - Site Location",
        "Check content": "LOG - Ghi log khi Download dữ liệu site location THẤT BẠI",
        "Pre-condition / Test Data": "Đang ở tab site_location.\nOnline nhưng server trả lỗi khi download (giả lập lỗi API).",
        "Steps": "1. Click nút ダウンロード\n2. Chọn YES\n3. Server trả về lỗi\n4. Toast báo lỗi hiển thị\n5. Kiểm tra log hệ thống",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} download data inspection result failed for site_location_id:{site_location_id} inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Log FAILED (không phải successful) được ghi\n- Chứa đúng site_location_id",
        "Priority": "High",
    },

    # L04 - Upload site location thành công
    {
        "TC No.": "MINSP_005_LOG_004",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Upload - Site Location",
        "Check content": "LOG - Ghi log khi Upload dữ liệu site location THÀNH CÔNG",
        "Pre-condition / Test Data": "Đang ở tab site_location.\nĐang Online.\nCó data local để upload.",
        "Steps": "1. Click nút アップロード (Upload) tại tab site_location\n2. Chọn YES tại popup xác nhận\n3. Upload thành công\n4. Kiểm tra log hệ thống",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} upload data inspection result successful for site_location_id:{site_location_id} inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Chứa đúng site_location_id của phòng đang upload\n- Đầy đủ các trường bắt buộc",
        "Priority": "High",
    },

    # L05 - Upload site location thất bại
    {
        "TC No.": "MINSP_005_LOG_005",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Upload - Site Location",
        "Check content": "LOG - Ghi log khi Upload dữ liệu site location THẤT BẠI",
        "Pre-condition / Test Data": "Đang ở tab site_location.\nOnline nhưng server lỗi khi nhận upload.",
        "Steps": "1. Click nút アップロード\n2. Chọn YES\n3. Server trả về lỗi\n4. Toast báo lỗi hiển thị\n5. Kiểm tra log hệ thống",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} upload data inspection result failed for site_location_id:{site_location_id} inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Log FAILED được ghi\n- site_location_id chính xác",
        "Priority": "High",
    },

    # L06 - Download Master Room thành công
    {
        "TC No.": "MINSP_005_LOG_006",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Download - Master Room",
        "Check content": "LOG - Ghi log khi Download dữ liệu Master Room THÀNH CÔNG",
        "Pre-condition / Test Data": "Đang ở tab Master Room (取引電力計).\nĐang Online.",
        "Steps": "1. Click nút ダウンロード tại Master Room\n2. Chọn YES\n3. Download thành công\n4. Kiểm tra log hệ thống",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} download data inspection result successful for master room inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Log chứa 'master room' (không có site_location_id)\n- Đầy đủ các trường bắt buộc",
        "Priority": "High",
    },

    # L07 - Download Master Room thất bại
    {
        "TC No.": "MINSP_005_LOG_007",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Download - Master Room",
        "Check content": "LOG - Ghi log khi Download dữ liệu Master Room THẤT BẠI",
        "Pre-condition / Test Data": "Đang ở Master Room. Online nhưng server lỗi.",
        "Steps": "1. Click ダウンロード\n2. Chọn YES\n3. Server trả lỗi\n4. Kiểm tra log",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} download data inspection result failed for master room inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Log FAILED cho master room\n- Không có site_location_id trong log",
        "Priority": "High",
    },

    # L08 - Upload Master Room thành công
    {
        "TC No.": "MINSP_005_LOG_008",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Upload - Master Room",
        "Check content": "LOG - Ghi log khi Upload dữ liệu Master Room THÀNH CÔNG",
        "Pre-condition / Test Data": "Đang ở Master Room. Đang Online. Có data local.",
        "Steps": "1. Click アップロード tại Master Room\n2. Chọn YES\n3. Upload thành công\n4. Kiểm tra log",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} upload data inspection result successful for master room inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Log chứa 'master room'\n- Đầy đủ 5 trường bắt buộc",
        "Priority": "High",
    },

    # L09 - Upload Master Room thất bại
    {
        "TC No.": "MINSP_005_LOG_009",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Upload - Master Room",
        "Check content": "LOG - Ghi log khi Upload dữ liệu Master Room THẤT BẠI",
        "Pre-condition / Test Data": "Đang ở Master Room. Online nhưng server lỗi upload.",
        "Steps": "1. Click アップロード\n2. Chọn YES\n3. Server trả lỗi\n4. Kiểm tra log",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} upload data inspection result failed for master room inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Log FAILED cho master room",
        "Priority": "High",
    },

    # L10 - Submit inspection
    {
        "TC No.": "MINSP_005_LOG_010",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Submit Inspection",
        "Check content": "LOG - Ghi log khi click nút Submit (顧客確認へ)",
        "Pre-condition / Test Data": "Đã nhập đủ required fields.\nĐang Online.",
        "Steps": "1. Click nút 顧客確認へ\n2. Hệ thống xử lý submit\n3. Kiểm tra log hệ thống",
        "Expected Result": "Log được ghi với format:\nUser {logged_in_user_id} submit inspection. inspection_id={id} {access_url} {date_time} {ip_address} {user_agent}\n\n- Log được ghi NGAY KHI bấm nút (dù submit thành công hay thất bại)\n- inspection_id và user_id đúng với inspection đang thực hiện",
        "Priority": "Critical",
    },

    # L11 - Kiểm tra KHÔNG ghi log khi Offline
    {
        "TC No.": "MINSP_005_LOG_011",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Offline - No Log",
        "Check content": "LOG - Không ghi log khi thực hiện offline (Download/Upload bị block bởi toast)",
        "Pre-condition / Test Data": "Tắt mạng (Offline).",
        "Steps": "1. Tắt mạng\n2. Click ダウンロード hoặc アップロード\n3. Toast báo lỗi offline hiện ra\n4. Kiểm tra log hệ thống",
        "Expected Result": "KHÔNG có log nào được ghi cho action Download/Upload khi offline vì action không thực sự được thực thi",
        "Priority": "Medium",
    },

    # L12 - Log khi Download bị cancel (chọn NO)
    {
        "TC No.": "MINSP_005_LOG_012",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Cancel Download/Upload",
        "Check content": "LOG - Không ghi log khi hủy Download/Upload (chọn NO)",
        "Pre-condition / Test Data": "Đang Online, có popup xác nhận.",
        "Steps": "1. Click ダウンロード hoặc アップロード\n2. Popup hiển thị\n3. Chọn NO\n4. Kiểm tra log",
        "Expected Result": "KHÔNG có log nào được ghi vì action bị cancel (không thực sự thực hiện)",
        "Priority": "Medium",
    },

    # L13 - Kiểm tra đầy đủ các trường trong log
    {
        "TC No.": "MINSP_005_LOG_013",
        "Check Object 1": "S09_Log",
        "Check Object 2": "Log Format Validation",
        "Check content": "LOG - Kiểm tra đầy đủ 5 trường bắt buộc trong mọi log entry",
        "Pre-condition / Test Data": "Thực hiện các action ghi log (access, download, upload, submit).",
        "Steps": "1. Thực hiện lần lượt: vào master room, download, upload, submit\n2. Lấy log từ hệ thống\n3. Kiểm tra từng trường của log entry",
        "Expected Result": "Mỗi log entry PHẢI có đầy đủ:\n1. logged_in_user_id - ID user đang đăng nhập\n2. inspection_id - ID của inspection\n3. access_url - URL đang truy cập\n4. date_time - Thời gian thực hiện\n5. ip_address - IP của thiết bị\n6. user_agent - Thông tin thiết bị/browser\n\nKhông có trường nào được NULL hoặc blank",
        "Priority": "High",
    },
]

# ============================================================
# APPEND Log cases vào file Excel hiện có
# ============================================================
output_path = r'c:\Users\Dell 5540\Downloads\antigravity-testing-kit\testcases_MINSP_005_FULL.xlsx'

wb = load_workbook(output_path)
ws = wb["MOB-010 Inspection Execute"]

# Style definitions
critical_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
high_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
medium_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
log_group_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
border = Border(
    left=Side(style='thin', color="D0D0D0"),
    right=Side(style='thin', color="D0D0D0"),
    top=Side(style='thin', color="D0D0D0"),
    bottom=Side(style='thin', color="D0D0D0")
)

# Find next empty row
next_row = ws.max_row + 1

# Add separator row
ws.cell(row=next_row, column=1, value="")
sep_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
sep_font = Font(color="FFFFFF", bold=True, size=11)
ws.cell(row=next_row, column=2, value="S09_Log - Kiểm tra Log System (Section 7. Log trong BD)")
ws.cell(row=next_row, column=2).fill = sep_fill
ws.cell(row=next_row, column=2).font = sep_font
ws.cell(row=next_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[next_row].height = 30
next_row += 1

# Column order must match existing headers
col_order = ["TC No.", "Check Object 1", "Check Object 2", "Check content",
             "Pre-condition / Test Data", "Steps", "Expected Result", "Priority"]

for case in log_cases:
    priority = case.get("Priority", "Medium")
    for col_idx, col_name in enumerate(col_order, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=case.get(col_name, ""))
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = border

        if priority == "Critical":
            cell.fill = critical_fill
        elif priority == "High":
            cell.fill = high_fill
        else:
            cell.fill = medium_fill

    ws.row_dimensions[next_row].height = 100
    next_row += 1

# Update Summary sheet
ws_sum = wb["Summary"]
# Find total row and update
existing_total_row = None
for row in ws_sum.iter_rows():
    for cell in row:
        if cell.value == "TOTAL":
            existing_total_row = cell.row
            # Update count in column B
            ws_sum.cell(row=existing_total_row, column=2).value = (
                ws_sum.cell(row=existing_total_row, column=2).value or 0
            ) + len(log_cases)
            break

# Add log summary info
last_sum_row = ws_sum.max_row + 2
ws_sum.cell(row=last_sum_row, column=1, value="Log Test Cases")
ws_sum.cell(row=last_sum_row, column=2, value=len(log_cases))
ws_sum.cell(row=last_sum_row, column=1).font = Font(bold=True)

wb.save(output_path)
print(f"Added {len(log_cases)} Log test cases to: {output_path}")
print(f"New total rows: {ws.max_row - 1} (excluding header)")
