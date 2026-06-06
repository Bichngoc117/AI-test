import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# DATA DICTIONARY: Tất cả các fields của màn hình MOB-010
# ============================================================

# Format: {field_id, label, group1, group2, data_type, required, min, max, format_desc, auto_calc}
MASTER_ROOM_FIELDS = [
    # Header Fields
    {"id": "back_btn",                          "label": "Back Button",           "g1": "S01_Header",   "g2": "Navigation",        "type": "button",  "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "inspection_category",               "label": "inspection_category",   "g1": "S01_Header",   "g2": "Thông tin Header",  "type": "label",   "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "plan_title",                        "label": "plan_title",            "g1": "S01_Header",   "g2": "Thông tin Header",  "type": "label",   "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "work_start",                        "label": "work_start",            "g1": "S01_Header",   "g2": "Thông tin Header",  "type": "label",   "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "execute_incharge",                  "label": "担当者",                 "g1": "S01_Header",   "g2": "Thông tin Header",  "type": "label",   "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "weather_temperature_humidity",      "label": "Thời tiết",             "g1": "S01_Header",   "g2": "Thông tin Header",  "type": "label",   "required": False, "min": None, "max": None, "fmt": None,   "calc": False},

    # Master Room - Sidebar
    {"id": "sidebar_expand",                    "label": "Sidebar Mở rộng",       "g1": "S02_Sidebar",  "g2": "Sidebar",           "type": "button",  "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "sidebar_collapse",                  "label": "Sidebar Thu gọn",       "g1": "S02_Sidebar",  "g2": "Sidebar",           "type": "button",  "required": False, "min": None, "max": None, "fmt": None,   "calc": False},

    # Master Room Tab Bar
    {"id": "site_location_area",                "label": "Tab Phòng",             "g1": "S03_Tab",      "g2": "Tab Bar",           "type": "tab",     "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "complete_icon",                     "label": "Icon Hoàn thành",       "g1": "S03_Tab",      "g2": "Tab Bar",           "type": "icon",    "required": False, "min": None, "max": None, "fmt": None,   "calc": False},

    # Sync buttons
    {"id": "sync_download_btn",                 "label": "Download Btn",          "g1": "S04_Sync",     "g2": "Đồng bộ dữ liệu",  "type": "button",  "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "sync_upload_btn",                   "label": "Upload Btn",            "g1": "S04_Sync",     "g2": "Đồng bộ dữ liệu",  "type": "button",  "required": False, "min": None, "max": None, "fmt": None,   "calc": False},
    {"id": "sync_datetime",                     "label": "Thời gian đồng bộ",     "g1": "S04_Sync",     "g2": "Đồng bộ dữ liệu",  "type": "label",   "required": False, "min": None, "max": None, "fmt": None,   "calc": False},

    # ① 受電状況 - Master Room Required Fields
    {"id": "meter_reading_date",                "label": "検針日",                 "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "date",    "required": True,  "min": None, "max": "today", "fmt": "MM/DD", "calc": False},
    {"id": "chk_hv_panel_demand_setpoint_num",  "label": "デマンド指示値",          "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": True,  "min": 0,    "max": 99999,   "fmt": "(5,3)", "calc": False},
    {"id": "chk_hv_panel_power_factor_rate",    "label": "力率",                   "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": False, "min": None, "max": None,    "fmt": "integer","calc": True},
    {"id": "chk_meter_replaced_flag",           "label": "メーター交換",            "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "boolean", "required": True,  "min": None, "max": None,    "fmt": "なし/あり","calc": False},
    {"id": "chk_energy_usage1_num",             "label": "使用電力量1",             "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": True,  "min": 0,    "max": 999999,  "fmt": "(6,3)", "calc": False},
    {"id": "chk_active_energy_p1_num",          "label": "有効電力量P1",            "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": True,  "min": 0,    "max": 999999,  "fmt": "(6,3)", "calc": False},
    {"id": "chk_reactive_energy_p2_num",        "label": "無効電力量P2",            "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": True,  "min": 0,    "max": 999999,  "fmt": "(6,3)", "calc": False},
    {"id": "chk_energy_usage2_num",             "label": "使用電力量2",             "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": False, "min": 0,    "max": 999999,  "fmt": "(6,3)", "calc": False},
    {"id": "chk_energy_usage3_num",             "label": "使用電力量3",             "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": False, "min": 0,    "max": 999999,  "fmt": "(6,3)", "calc": False},
    {"id": "total_usage",                       "label": "総使用量",               "g1": "S05_Step1_受電状況", "g2": "受電状況",       "type": "decimal", "required": False, "min": None, "max": None,    "fmt": "auto",  "calc": True},

    # ② 高圧受電盤 - Master Room Step 2
    {"id": "chk_voltage_rs_num_master",         "label": "電圧(V) R-S",            "g1": "S06_Step2_高圧受電盤", "g2": "高圧受電盤",   "type": "integer", "required": False, "min": 5000, "max": 8000,    "fmt": None,   "calc": False},
    {"id": "chk_voltage_st_num_master",         "label": "電圧(V) S-T",            "g1": "S06_Step2_高圧受電盤", "g2": "高圧受電盤",   "type": "integer", "required": False, "min": 5000, "max": 8000,    "fmt": None,   "calc": False},
    {"id": "chk_voltage_tr_num_master",         "label": "電圧(V) T-R",            "g1": "S06_Step2_高圧受電盤", "g2": "高圧受電盤",   "type": "integer", "required": False, "min": 5000, "max": 8000,    "fmt": None,   "calc": False},
    {"id": "chk_amp_r_num_master",              "label": "電流(A) R",              "g1": "S06_Step2_高圧受電盤", "g2": "高圧受電盤",   "type": "decimal", "required": False, "min": 0,    "max": 600,     "fmt": "(3,2)", "calc": False},
    {"id": "chk_amp_s_num_master",              "label": "電流(A) S",              "g1": "S06_Step2_高圧受電盤", "g2": "高圧受電盤",   "type": "decimal", "required": False, "min": 0,    "max": 600,     "fmt": "(3,2)", "calc": False},
    {"id": "chk_amp_t_num_master",              "label": "電流(A) T",              "g1": "S06_Step2_高圧受電盤", "g2": "高圧受電盤",   "type": "decimal", "required": False, "min": 0,    "max": 600,     "fmt": "(3,2)", "calc": False},

    # ③ 測定器 - Borrowed Equipment
    {"id": "borrowed_equipment_section",        "label": "測定器・試験器一覧表",    "g1": "S07_Step3_測定器",   "g2": "測定器リスト",   "type": "section", "required": False, "min": None, "max": None,    "fmt": None,   "calc": False},
]

SITE_LOCATION_FIELDS = [
    # Site Location - 受電設備 (Transformer)
    {"id": "transformer_chk_temperature_num",   "label": "温度(℃) - Transformer",  "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "decimal", "required": True,  "min": 0,    "max": 200,   "fmt": "(3,1)", "calc": False},
    {"id": "transformer_chk_load_factor_rate",  "label": "負荷率 - Transformer",   "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "decimal", "required": False, "min": 0,    "max": 100,   "fmt": "(3,2)", "calc": False},
    {"id": "transformer_chk_leakage_current",   "label": "漏れ電流 - Transformer", "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "decimal", "required": True,  "min": 0,    "max": 9999,  "fmt": "(4,1)", "calc": False},
    {"id": "transformer_chk_voltage_rs",        "label": "電圧(V) R-S - Trans.",   "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "integer", "required": True,  "min": 50,   "max": 3600,  "fmt": None,   "calc": False},
    {"id": "transformer_chk_voltage_st",        "label": "電圧(V) S-T - Trans.",   "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "integer", "required": True,  "min": 50,   "max": 3600,  "fmt": None,   "calc": False},
    {"id": "transformer_chk_voltage_tr",        "label": "電圧(V) T-R - Trans.",   "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "integer", "required": True,  "min": 50,   "max": 3600,  "fmt": None,   "calc": False},
    {"id": "transformer_chk_amp_r",             "label": "電流(A) R - Trans.",     "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "decimal", "required": True,  "min": 0,    "max": 3000,  "fmt": "(4,1)", "calc": False},
    {"id": "transformer_chk_amp_s",             "label": "電流(A) S - Trans.",     "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "decimal", "required": True,  "min": 0,    "max": 3000,  "fmt": "(4,1)", "calc": False},
    {"id": "transformer_chk_amp_t",             "label": "電流(A) T - Trans.",     "g1": "SIT_受電設備",   "g2": "変圧器負荷状況",  "type": "decimal", "required": True,  "min": 0,    "max": 3000,  "fmt": "(4,1)", "calc": False},

    # その他高圧機器 (Condenser/Reactor)
    {"id": "condenser_chk_temperature_num",     "label": "温度(℃) - Other HV",    "g1": "SIT_受電設備",   "g2": "その他高圧機器",  "type": "decimal", "required": True,  "min": 0,    "max": 200,   "fmt": "(3,1)", "calc": False},

    # 絶縁監視装置 (Insulation Monitoring)
    {"id": "insulation_chk_communication_test", "label": "通信テスト",             "g1": "SIT_絶縁監視",   "g2": "絶縁監視装置",    "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},
    {"id": "insulation_chk_alarm_status",       "label": "警報の状況",             "g1": "SIT_絶縁監視",   "g2": "絶縁監視装置",    "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},
    {"id": "insulation_ch1_measured_value",     "label": "CH1 測定値",            "g1": "SIT_絶縁監視",   "g2": "絶縁監視装置",    "type": "decimal", "required": False, "min": None, "max": None,  "fmt": None,   "calc": False},
    {"id": "insulation_ch2_measured_value",     "label": "CH2 測定値",            "g1": "SIT_絶縁監視",   "g2": "絶縁監視装置",    "type": "decimal", "required": False, "min": None, "max": None,  "fmt": None,   "calc": False},
    {"id": "insulation_ch3_measured_value",     "label": "CH3 測定値",            "g1": "SIT_絶縁監視",   "g2": "絶縁監視装置",    "type": "decimal", "required": False, "min": None, "max": None,  "fmt": None,   "calc": False},
    {"id": "insulation_ch4_measured_value",     "label": "CH4 測定値",            "g1": "SIT_絶縁監視",   "g2": "絶縁監視装置",    "type": "decimal", "required": False, "min": None, "max": None,  "fmt": None,   "calc": False},
    {"id": "insulation_overall_judgment",       "label": "総合判定",              "g1": "SIT_絶縁監視",   "g2": "絶縁監視装置",    "type": "select",  "required": False, "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 蓄電設備 (Storage Equipment)
    {"id": "storage_chk_select_status",         "label": "蓄電 Status",           "g1": "SIT_蓄電設備",   "g2": "蓄電設備",        "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 設備点検結果 (Inspection Location)
    {"id": "location_chk_overall",              "label": "設備点検結果 Overall",   "g1": "SIT_設備点検結果","g2": "設備点検結果",    "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},
    {"id": "location_chk_reason",               "label": "設備点検結果 Reason",    "g1": "SIT_設備点検結果","g2": "設備点検結果",    "type": "text",    "required": False, "min": None, "max": None,  "fmt": None,   "calc": False},

    # 非常用・常用発電機 (Generator)
    {"id": "generator_chk_select",              "label": "発電機 Status",         "g1": "SIT_発電機",     "g2": "発電機",          "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 電気室キュービクル (Cubicle)
    {"id": "cubicle_chk_select",                "label": "電気室 Status",         "g1": "SIT_キュービクル","g2": "電気室キュービクル","type": "select", "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 遮断器 (Switchgear)
    {"id": "switchgear_chk_select",             "label": "遮断器 Status",         "g1": "SIT_遮断器",     "g2": "遮断器・断路器",   "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 保護継電器 - Ground
    {"id": "relay_ground_op_time",              "label": "動作時間 - Ground",     "g1": "SIT_保護継電器", "g2": "高圧地絡継電器",   "type": "decimal", "required": True,  "min": 0,    "max": 9999,  "fmt": "(4,2)", "calc": False},
    {"id": "relay_ground_measured_value",       "label": "測定値 - Ground",       "g1": "SIT_保護継電器", "g2": "高圧地絡継電器",   "type": "decimal", "required": False, "min": 0,    "max": 9999,  "fmt": "(4,2)", "calc": False},
    {"id": "relay_ground_result",               "label": "結果 - Ground",         "g1": "SIT_保護継電器", "g2": "高圧地絡継電器",   "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 保護継電器 - Overcurrent
    {"id": "relay_overcurrent_op_time",         "label": "動作時間 - OC",         "g1": "SIT_保護継電器", "g2": "過電流継電器",     "type": "decimal", "required": True,  "min": 0,    "max": 9999,  "fmt": "(4,2)", "calc": False},
    {"id": "relay_overcurrent_measured",        "label": "測定値 - OC",           "g1": "SIT_保護継電器", "g2": "過電流継電器",     "type": "decimal", "required": False, "min": 0,    "max": 9999,  "fmt": "(4,2)", "calc": False},
    {"id": "relay_overcurrent_result",          "label": "結果 - OC",             "g1": "SIT_保護継電器", "g2": "過電流継電器",     "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 保護継電器 - Voltage
    {"id": "relay_voltage_result",              "label": "電圧継電器 Result",      "g1": "SIT_保護継電器", "g2": "電圧継電器",       "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 接地抵抗測定 (Ground Resistance)
    {"id": "ground_resistance_measured",        "label": "接地抵抗測定値",         "g1": "SIT_接地測定",   "g2": "接地抵抗測定",     "type": "decimal", "required": True,  "min": 0,    "max": 9999,  "fmt": "(4,2)", "calc": False},
    {"id": "ground_resistance_result",          "label": "接地抵抗結果",           "g1": "SIT_接地測定",   "g2": "接地抵抗測定",     "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 高圧絶縁抵抗 (HV Insulation Resistance)
    {"id": "hv_insulation_measured",            "label": "高圧絶縁抵抗値",         "g1": "SIT_絶縁抵抗",   "g2": "高圧絶縁抵抗測定", "type": "decimal", "required": True,  "min": 0,    "max": 9999,  "fmt": "(4,2)", "calc": False},
    {"id": "hv_insulation_result",              "label": "高圧絶縁抵抗結果",       "g1": "SIT_絶縁抵抗",   "g2": "高圧絶縁抵抗測定", "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},

    # 低圧絶縁抵抗 (LV Insulation Resistance)
    {"id": "lv_insulation_measured",            "label": "低圧絶縁抵抗値",         "g1": "SIT_絶縁抵抗",   "g2": "低圧絶縁抵抗測定", "type": "decimal", "required": True,  "min": 0,    "max": 9999,  "fmt": "(4,2)", "calc": False},
    {"id": "lv_insulation_result",              "label": "低圧絶縁抵抗結果",       "g1": "SIT_絶縁抵抗",   "g2": "低圧絶縁抵抗測定", "type": "select",  "required": True,  "min": None, "max": None,  "fmt": "◯△×ー","calc": False},
]

# ============================================================
# TEST CASE RULE ENGINE
# ============================================================
def generate_cases_for_field(field, tc_counter):
    cases = []
    fid = field["id"]
    label = field["label"]
    g1 = field["g1"]
    g2 = field["g2"]
    dtype = field["type"]
    required = field["required"]
    fmin = field["min"]
    fmax = field["max"]
    fmt = field.get("fmt")
    calc = field.get("calc", False)

    def tc(num, obj2, content, precond, steps, expected, priority="Medium"):
        return {
            "TC No.": f"MINSP_005_{str(num).zfill(3)}",
            "Check Object 1": g1,
            "Check Object 2": obj2,
            "Check content": content,
            "Pre-condition / Test Data": precond,
            "Steps": steps,
            "Expected Result": expected,
            "Priority": priority,
        }

    n = tc_counter[0]

    # --- Numeric / Decimal / Integer fields ---
    if dtype in ("decimal", "integer"):
        # Happy Path - EP Valid
        if fmin is not None and fmax is not None:
            valid_val = (fmin + fmax) / 2
            if dtype == "integer":
                valid_val = int(valid_val)
            cases.append(tc(n, f"{label}", "EP - Nhập giá trị hợp lệ",
                f"Đang ở màn hình có trường {label}",
                f"1. Tap vào trường {label}\n2. Nhập {valid_val}\n3. Focus out",
                f"1. Bàn phím số hiển thị\n2. Giá trị {valid_val} được hiển thị\n3. Lưu thành công, background đổi vàng nếu có thay đổi",
                "High")); n += 1

        # BVA - Boundary Min
        if fmin is not None:
            cases.append(tc(n, f"{label}", "BVA - Nhập đúng Min",
                f"Đang ở trường {label}",
                f"1. Nhập {fmin}\n2. Focus out",
                f"1. Giá trị {fmin} lưu thành công\n2. Không báo lỗi",
                "High")); n += 1

            cases.append(tc(n, f"{label}", "BVA - Nhập dưới Min",
                f"Đang ở trường {label}",
                f"1. Nhập {fmin - 1}\n2. Focus out",
                f"1. Báo lỗi 'Input range'\n2. Clear giá trị đã nhập",
                "High")); n += 1

        # BVA - Boundary Max
        if fmax is not None and isinstance(fmax, (int, float)):
            cases.append(tc(n, f"{label}", "BVA - Nhập đúng Max",
                f"Đang ở trường {label}",
                f"1. Nhập {fmax}\n2. Focus out",
                f"1. Giá trị {fmax} lưu thành công\n2. Không báo lỗi",
                "High")); n += 1

            cases.append(tc(n, f"{label}", "BVA - Nhập vượt Max",
                f"Đang ở trường {label}",
                f"1. Nhập {fmax + 1}\n2. Focus out",
                f"1. Báo lỗi 'Input range'\n2. Clear giá trị đã nhập",
                "High")); n += 1

        # EP - Negative number (invalid)
        if fmin is not None and fmin >= 0:
            cases.append(tc(n, f"{label}", "EP - Nhập số âm",
                f"Đang ở trường {label}",
                f"1. Nhập -1\n2. Focus out",
                f"1. Báo lỗi 'Input range'\n2. Clear giá trị đã nhập",
                "Medium")); n += 1

        # Required check
        if required:
            cases.append(tc(n, f"{label}", "Validation - Bỏ trống (Required)",
                f"Đang ở màn hình có trường {label}",
                f"1. Để trống trường {label}\n2. Focus out",
                f"1. Hiển thị error 'Required' (viền đỏ)\n2. Không submit được",
                "Critical")); n += 1

        # Auto-calc check
        if calc:
            cases.append(tc(n, f"{label}", "Auto Calculate - Kiểm tra công thức",
                f"Đã nhập đủ các input cần thiết để tính {label}",
                f"1. Nhập giá trị cho các field liên quan\n2. Kiểm tra {label}",
                f"1. {label} được tự động tính theo công thức\n2. Làm tròn theo đúng format",
                "High")); n += 1

            cases.append(tc(n, f"{label}", "Auto Calculate - Xử lý edge case (0/0)",
                f"Nhập giá trị có thể gây chia cho 0",
                f"1. Nhập các input = 0\n2. Kiểm tra {label}",
                f"1. Không crash\n2. Hiển thị 0 hoặc N/A theo spec",
                "High")); n += 1

        # UI Highlight - background vàng
        if not calc:
            cases.append(tc(n, f"{label}", "UI - Background vàng khi có thay đổi",
                f"Đã có data last time cho {label}",
                f"1. Nhập giá trị khác với last time\n2. Quan sát UI",
                f"1. Background của field chuyển sang vàng\n2. Last time vẫn hiển thị bên cạnh",
                "Medium")); n += 1

    # --- Date field ---
    elif dtype == "date":
        cases.append(tc(n, f"{label}", "Happy Path - Chọn ngày hợp lệ",
            f"Đang ở trường {label}",
            f"1. Tap vào trường {label}\n2. Calendar mở ra\n3. Chọn ngày hôm nay (today)",
            f"1. Calendar hiển thị\n2. Ngày được chọn và lưu vào field theo format MM/DD",
            "High")); n += 1

        cases.append(tc(n, f"{label}", "BVA - Không cho chọn ngày tương lai",
            f"Đang ở Calendar",
            f"1. Mở Calendar\n2. Thử chọn ngày mai",
            f"1. Ngày mai bị disable, không thể chọn",
            "High")); n += 1

        cases.append(tc(n, f"{label}", "BVA - Không cho chọn ngày <= last_time",
            f"Có last_time = 04/15",
            f"1. Mở Calendar\n2. Thử chọn ngày 04/10",
            f"1. Ngày 04/10 bị disable, không thể chọn",
            "High")); n += 1

        cases.append(tc(n, f"{label}", "Auto-fill - Có last_time",
            f"Last_time = 03/15, tháng hiện tại = 04",
            f"1. Vào màn hình lần đầu\n2. Kiểm tra {label}",
            f"1. Tự động điền 04/15 (giữ nguyên ngày, thay tháng)",
            "High")); n += 1

        if required:
            cases.append(tc(n, f"{label}", "Validation - Bỏ trống (Required)",
                f"Không có last_time",
                f"1. Để trống {label}\n2. Focus out",
                f"1. Hiển thị error 'Required' (viền đỏ)",
                "Critical")); n += 1

        cases.append(tc(n, f"{label}", "UI - Background vàng khi sửa",
            f"Đã có last_time",
            f"1. Sửa ngày khác với last_time\n2. Quan sát",
            f"1. Background chuyển vàng",
            "Medium")); n += 1

    # --- Boolean field ---
    elif dtype == "boolean":
        cases.append(tc(n, f"{label}", "Happy Path - Chọn TRUE (あり)",
            f"Đang ở màn hình có trường {label}",
            f"1. Tap 'あり'\n2. Kiểm tra UI",
            f"1. 'あり' có background xanh, text trắng\n2. 'なし' background xám",
            "High")); n += 1

        cases.append(tc(n, f"{label}", "Happy Path - Chọn FALSE (なし)",
            f"Đang ở màn hình có trường {label}",
            f"1. Tap 'なし'\n2. Kiểm tra UI",
            f"1. 'なし' có background xanh, text trắng\n2. 'あり' background xám",
            "High")); n += 1

        cases.append(tc(n, f"{label}", "Default State - Giá trị mặc định",
            f"Mở màn hình lần đầu",
            f"1. Kiểm tra {label} khi chưa tương tác",
            f"1. Mặc định chọn 'なし' (false)",
            "Medium")); n += 1

    # --- Select field (◯△×ー) ---
    elif dtype == "select":
        for option, meaning in [("◯", "IS_GOOD(2)"), ("△", "IS_WARNING(1)"), ("×", "IS_BAD(-1)"), ("ー", "IS_NOT_APPLICABLE(0)")]:
            cases.append(tc(n, f"{label}", f"Happy Path - Chọn '{option}' ({meaning})",
                f"Đang ở trường {label}",
                f"1. Tap vào option '{option}'\n2. Kiểm tra UI",
                f"1. '{option}' được highlight\n2. Giá trị {meaning} được lưu",
                "High")); n += 1

        if required:
            cases.append(tc(n, f"{label}", "Validation - Chưa chọn option (Required)",
                f"Chưa tương tác với {label}",
                f"1. Để mặc định không chọn\n2. Thử submit",
                f"1. Báo lỗi required\n2. Không submit được",
                "Critical")); n += 1

    # --- Button field ---
    elif dtype == "button":
        if "Download" in label or "Upload" in label:
            for net, state in [("Online", "bật mạng"), ("Offline", "tắt mạng")]:
                cases.append(tc(n, f"{label}", f"Click khi {state} ({net})",
                    f"Trạng thái mạng: {state}",
                    f"1. {state.capitalize()}\n2. Click {label}",
                    f"{'Hiện popup xác nhận YES/NO' if net == 'Online' else 'Hiện Toast báo lỗi offline'}",
                    "High")); n += 1
            # Click YES/NO
            cases.append(tc(n, f"{label}", "Click Online -> Chọn YES",
                f"Có mạng",
                f"1. Click {label}\n2. Popup hiện\n3. Chọn YES",
                f"1. Thực hiện {label} action thành công\n2. Hiện Toast thành công\n3. Cập nhật sync_datetime",
                "Critical")); n += 1
            cases.append(tc(n, f"{label}", "Click Online -> Chọn NO",
                f"Có mạng",
                f"1. Click {label}\n2. Popup hiện\n3. Chọn NO",
                f"1. Đóng popup\n2. Không thực hiện action",
                "Low")); n += 1

        elif "Sidebar" in label or "expand" in label.lower() or "collapse" in label.lower():
            cases.append(tc(n, f"{label}", "Toggle Sidebar",
                f"Sidebar ở trạng thái hiện tại",
                f"1. Click {label}",
                f"1. Sidebar chuyển trạng thái (mở <-> thu gọn)",
                "Low")); n += 1

        elif "back" in fid.lower():
            cases.append(tc(n, f"{label}", "Click Back - Quay lại màn trước",
                f"Đang ở MOB-010",
                f"1. Click nút Back\n2. Kiểm tra màn hình",
                f"1. Quay lại màn hình trước đó (MOB-002 hoặc MOB-008)\n2. Data được làm mới",
                "Medium")); n += 1
        else:
            cases.append(tc(n, f"{label}", f"Click {label}",
                f"Điều kiện bình thường",
                f"1. Click {label}",
                f"1. Thực hiện đúng chức năng của {label}",
                "Medium")); n += 1

    # --- Label field ---
    elif dtype == "label":
        cases.append(tc(n, f"{label}", "Hiển thị đúng dữ liệu",
            f"Có data hợp lệ trong DB",
            f"1. Vào màn hình\n2. Kiểm tra {label}",
            f"1. {label} hiển thị đúng giá trị từ DB",
            "Medium")); n += 1

        cases.append(tc(n, f"{label}", "Không có data",
            f"Không có data / NULL",
            f"1. Vào màn hình với data = NULL\n2. Kiểm tra {label}",
            f"1. {label} hiển thị blank hoặc '-' theo spec",
            "Low")); n += 1

    # --- Tab field ---
    elif dtype == "tab":
        cases.append(tc(n, f"{label}", "Chuyển Tab - Phòng khác",
            f"Có nhiều phòng (tabs)",
            f"1. Tap vào tab phòng khác",
            f"1. Nội dung đổi sang phòng đã chọn\n2. Tab được highlight",
            "Medium")); n += 1

        cases.append(tc(n, f"{label}", "Tab hoàn thành - UI trạng thái DONE",
            f"Đã nhập đủ required fields cho phòng",
            f"1. Nhập đủ data\n2. Quan sát tab UI",
            f"1. Tab có icon check\n2. Background xanh nhạt",
            "High")); n += 1

        cases.append(tc(n, f"{label}", "Scroll ngang khi nhiều tab",
            f"Có > 5 phòng",
            f"1. Cuộn ngang thanh tab\n2. Kiểm tra hiển thị",
            f"1. Tất cả tabs đều accessible bằng scroll ngang",
            "Low")); n += 1

    # --- Section field ---
    elif dtype == "section":
        cases.append(tc(n, f"{label}", "Ẩn section khi không có data",
            f"User không có thiết bị mượn",
            f"1. Vào màn hình\n2. Cuộn đến section {label}",
            f"1. Section {label} bị ẩn hoàn toàn",
            "Medium")); n += 1

        cases.append(tc(n, f"{label}", "Hiển thị đủ columns khi có data",
            f"User có 2 thiết bị mượn",
            f"1. Vào màn hình\n2. Cuộn đến section {label}",
            f"1. Hiển thị 2 rows với đủ cột: Tên,型式, Hãng, Ngày SX, Serial",
            "Medium")); n += 1

    tc_counter[0] = n
    return cases

# ============================================================
# MANUAL / FLOW-BASED TEST CASES
# ============================================================
def generate_flow_cases(start_n):
    n = start_n
    cases = []

    def tc(num, g1, g2, content, precond, steps, expected, priority="High"):
        return {
            "TC No.": f"MINSP_005_{str(num).zfill(3)}",
            "Check Object 1": g1,
            "Check Object 2": g2,
            "Check content": content,
            "Pre-condition / Test Data": precond,
            "Steps": steps,
            "Expected Result": expected,
            "Priority": priority,
        }

    # Navigation Flows
    cases.append(tc(n, "S00_Navigation", "Luồng tới màn hình", "Từ Home - Bấm 点検開始", "User đã login với role Engineer", "1. Open App\n2. Login role Engineer\n3. Tại MOB-002 Home, click '点検開始'", "1. Vào màn hình MOB-010 Inspection Execute\n2. Focus tab Master Room đầu tiên", "High")); n+=1
    cases.append(tc(n, "S00_Navigation", "Luồng tới màn hình", "Từ MOB-008 Plan Detail - Bấm 点検開始", "User đã login", "1. Click 詳細\n2. Tại MOB-008, click 点検開始", "1. Vào MOB-010", "High")); n+=1
    cases.append(tc(n, "S00_Navigation", "Luồng tới màn hình", "Từ MOB-011 Report - Bấm 点検項目を修正", "Có report đang soạn thảo", "1. Vào MOB-011\n2. Click '点検項目を修正'", "1. Vào MOB-010", "High")); n+=1
    cases.append(tc(n, "S00_Navigation", "Quay lại (点検再開)", "Đã nhập dở ở phòng B", "User đã nhập dở trước đó", "1. Thoát khỏi màn hình\n2. Quay lại bằng '点検再開' từ Home", "1. Vào đúng phòng B\n2. Dữ liệu đã nhập được giữ nguyên", "Critical")); n+=1
    cases.append(tc(n, "S00_Navigation", "Không có quyền truy cập", "User KHÔNG trong danh sách assign", "SQL trả về 0 record", "1. Thử truy cập plan\n2. Kiểm tra", "1. Redirect về MOB-002 HOME", "High")); n+=1
    cases.append(tc(n, "S00_Navigation", "Không có quyền truy cập", "SQL trả về 0 record", "Query điều kiện thất bại", "1. Truy cập với user không assign\n2. Kiểm tra", "1. Redirect về Home", "High")); n+=1
    cases.append(tc(n, "S00_Navigation", "Back Button - Refresh API", "Đang ở màn khác, bấm quay lại MOB-010", "Đang ở màn hình khác", "1. Bấm Back quay lại MOB-010", "1. Màn hình hiển thị lại\n2. Data NOT refresh (không reload)", "Medium")); n+=1

    # Completion flow
    cases.append(tc(n, "S08_Submit", "Complete Inspection", "Submit thành công - Tất cả phòng hoàn thành", "Đã nhập đủ required fields TẤT CẢ phòng", "1. Click '顧客確認へ'\n2. Đợi xử lý", "1. Validate pass\n2. Tạo Report ID dạng HJS0000000001\n3. Chuyển sang MOB-013", "Critical")); n+=1
    cases.append(tc(n, "S08_Submit", "Complete Inspection", "Submit thất bại - Thiếu required field ở 1 phòng", "Phòng B chưa nhập 1 field required", "1. Click '顧客確認へ'\n2. Kiểm tra", "1. Báo lỗi validation\n2. Focus vào field đang thiếu\n3. Không chuyển màn", "Critical")); n+=1
    cases.append(tc(n, "S08_Submit", "Complete Inspection", "Submit khi Offline", "Tắt Wifi", "1. Tắt mạng\n2. Click '顧客確認へ'", "1. Nút bấm được\n2. Hiện Toast báo offline\n3. Không submit", "High")); n+=1
    cases.append(tc(n, "S08_Submit", "Complete Inspection", "Lỗi insert Report vào DB", "Giả lập DB lỗi", "1. Submit hợp lệ\n2. DB báo lỗi khi insert report", "1. Hiện E-MSG-004\n2. Không chuyển màn\n3. User có thể click lại", "High")); n+=1
    cases.append(tc(n, "S08_Submit", "Complete Inspection", "Report đã tồn tại - Không tạo duplicate", "Inspection đã có report", "1. Submit lần 2 (sau khi sửa)", "1. Không insert record mới\n2. Skip insert, tiếp tục chuyển màn", "High")); n+=1

    # Execution Progress
    cases.append(tc(n, "S03_Tab", "execution_progress", "Hiển thị tiến độ hoàn thành phòng", "Hoàn thành 2/5 phòng", "1. Kiểm tra số hiển thị ở góc", "1. Hiển thị '2/5 完了'", "Medium")); n+=1
    cases.append(tc(n, "S03_Tab", "execution_progress", "Cập nhật khi phòng hoàn thành", "Vừa hoàn thành phòng thứ 3", "1. Nhập xong required fields phòng 3\n2. Kiểm tra counter", "1. Counter tự động update '3/5 完了'", "High")); n+=1

    # Multi-user collaboration
    cases.append(tc(n, "S04_Sync", "Multi-user scenario", "Nhiều người cùng thực hiện inspection", "2 kỹ thuật viên assign, A ở phòng 1, B ở phòng 2", "1. A nhập xong phòng 1, Upload\n2. B Download\n3. B kiểm tra dữ liệu phòng 1", "1. B thấy dữ liệu của A sau khi Download", "High")); n+=1
    cases.append(tc(n, "S04_Sync", "Multi-user scenario", "Ghi đè dữ liệu khi Download (có local thay đổi)", "Local có background vàng (đã sửa)", "1. Click Download\n2. Popup xuất hiện cảnh báo\n3. Chọn YES", "1. Dữ liệu local bị ghi đè\n2. Màu vàng mất (đồng bộ về server)", "High")); n+=1

    # Remark / Issue buttons
    cases.append(tc(n, "S01_Header", "remark_btn", "Mở màn thêm ghi chú (特記事項)", "Đang ở MOB-010", "1. Click 特記事項\n2. Kiểm tra", "1. Mở màn MOB-010-ADD-Remark", "Low")); n+=1
    cases.append(tc(n, "S01_Header", "remark_btn", "Số lượng ghi chú hiển thị trên button", "Có 3 ghi chú", "1. Kiểm tra nút 特記事項", "1. Hiển thị '特記事項 (3)'", "Low")); n+=1
    cases.append(tc(n, "S01_Header", "isue_btn", "Mở màn thêm chỉ trích (指摘事項)", "Đang ở MOB-010", "1. Click 指摘事項", "1. Mở màn MOB-010-ADD-Issue", "Low")); n+=1
    cases.append(tc(n, "S01_Header", "isue_btn", "Số lượng chỉ trích hiển thị trên button", "Có 2 chỉ trích", "1. Kiểm tra nút 指摘事項", "1. Hiển thị '指摘事項 (2)'", "Low")); n+=1

    # Modal Navigation
    cases.append(tc(n, "SIT_受電設備", "Modal Input", "Mở modal nhập liệu", "Đang ở tab site_location có 受電設備", "1. Tap vào field nhập số", "1. Modal mở\n2. Hiển thị đúng group đầu tiên (電圧)", "High")); n+=1
    cases.append(tc(n, "SIT_受電設備", "Modal Input", "Chuyển Group trong modal (vòng lặp)", "Trong modal group 電圧(V)", "1. Click '>\n2. Click '>\n3. Click '>'", "1. Group chuyển 電圧 -> 電流 -> 温度\n2. Click > lần nữa -> quay lại 電圧", "Medium")); n+=1
    cases.append(tc(n, "SIT_受電設備", "Modal Input", "Chuyển Equipment khi nhiều thiết bị", "Có 3 equipment", "1. Click Lên/Xuống", "1. Dữ liệu của thiết bị tương ứng được hiển thị", "Medium")); n+=1
    cases.append(tc(n, "SIT_受電設備", "Modal Input", "Đóng modal bằng cách click outside", "Trong modal", "1. Click vùng tối ngoài modal", "1. Modal đóng\n2. Dữ liệu đã nhập trong modal truyền ra bảng ngoài", "High")); n+=1
    cases.append(tc(n, "SIT_受電設備", "Modal Input", "Ẩn nút Lên/Xuống khi chỉ 1 equipment", "Chỉ có 1 equipment", "1. Kiểm tra modal", "1. Không hiển thị nút điều hướng thiết bị", "Low")); n+=1

    # Sidebar step completion
    cases.append(tc(n, "S02_Sidebar", "Step DONE State", "Step chuyển sang DONE sau khi nhập đủ", "Nhập đủ required fields của step 1", "1. Nhập đủ fields\n2. Quan sát sidebar", "1. Step có icon check\n2. Background xanh nhạt trong sidebar", "High")); n+=1
    cases.append(tc(n, "S02_Sidebar", "Room DONE State", "Phòng chuyển sang DONE sau tất cả steps DONE", "Tất cả steps hoàn thành", "1. Hoàn thành tất cả steps\n2. Quan sát tab phòng", "1. Tab phòng có icon check\n2. Background xanh nhạt", "Critical")); n+=1

    # Offline mode - general
    cases.append(tc(n, "SIT_Input", "Offline Mode", "Nhập data khi offline - Lưu local", "Tắt Wifi", "1. Tắt Wifi\n2. Nhập data vào các fields\n3. Quan sát", "1. Nhập data bình thường\n2. Tự động lưu vào local storage của thiết bị", "Critical")); n+=1
    cases.append(tc(n, "SIT_Input", "Offline Mode", "Thoát app và vào lại - Data còn đủ", "Đã nhập data offline rồi thoát app", "1. Nhập data\n2. Thoát app\n3. Mở app lại\n4. Vào lại inspection", "1. Data được giữ lại từ local storage", "Critical")); n+=1

    # Weather display
    cases.append(tc(n, "S01_Header", "weather_temperature_humidity", "Hiển thị icon thời tiết theo mapping", "Site có data thời tiết", "1. Kiểm tra icon thời tiết ở header", "1. Hiển thị đúng icon + nhiệt độ°C / độ ẩm%", "Low")); n+=1
    cases.append(tc(n, "S01_Header", "weather_temperature_humidity", "Không có data thời tiết", "Site không có data thời tiết", "1. Kiểm tra icon thời tiết", "1. Không crash\n2. Hiển thị blank hoặc '-'", "Low")); n+=1

    # Work Procedure button
    cases.append(tc(n, "S02_Sidebar", "work_procedure_btn", "Mở tài liệu hướng dẫn", "Đang ở màn hình", "1. Click icon 作業手順書", "1. Mở modal MOB-008 Inspect Work Procedure", "Low")); n+=1

    # Local storage per step
    cases.append(tc(n, "SIT_Input", "Auto-save local", "Tự động lưu sau mỗi lần nhập field", "Đang nhập trong step 2 của phòng A", "1. Nhập giá trị\n2. Focus out\n3. Tắt app đột ngột", "1. Khi vào lại, giá trị vừa nhập vẫn còn", "Critical")); n+=1

    return cases, n

# ============================================================
# MAIN: Tổng hợp và export
# ============================================================
all_cases = []
counter = [1]

all_fields = MASTER_ROOM_FIELDS + SITE_LOCATION_FIELDS

for field in all_fields:
    cases = generate_cases_for_field(field, counter)
    all_cases.extend(cases)

flow_cases, final_n = generate_flow_cases(counter[0])
all_cases.extend(flow_cases)

print(f"Total test cases generated: {len(all_cases)}")

df = pd.DataFrame(all_cases)

# ============================================================
# EXPORT TO EXCEL with styling
# ============================================================
output_path = r'c:\Users\Dell 5540\Downloads\antigravity-testing-kit\testcases_MINSP_005_FULL.xlsx'

wb = Workbook()
ws = wb.active
ws.title = "MOB-010 Inspection Execute"

# Header styling
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
alt_fill_1 = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
alt_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
critical_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
high_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
border = Border(
    left=Side(style='thin', color="D0D0D0"),
    right=Side(style='thin', color="D0D0D0"),
    top=Side(style='thin', color="D0D0D0"),
    bottom=Side(style='thin', color="D0D0D0")
)

columns = list(df.columns)
col_widths = {
    "TC No.": 18,
    "Check Object 1": 30,
    "Check Object 2": 30,
    "Check content": 40,
    "Pre-condition / Test Data": 45,
    "Steps": 60,
    "Expected Result": 60,
    "Priority": 12,
}

# Write header
for col_idx, col_name in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

ws.row_dimensions[1].height = 35

# Write data
for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
    row_list = list(row_data)
    priority = row_list[-1]

    for col_idx, value in enumerate(row_list, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = border

        # Priority-based row coloring
        if priority == "Critical":
            cell.fill = critical_fill
        elif priority == "High":
            cell.fill = high_fill
        elif row_idx % 2 == 0:
            cell.fill = alt_fill_1
        else:
            cell.fill = alt_fill_2

    ws.row_dimensions[row_idx].height = 70

# Freeze header row
ws.freeze_panes = "A2"

# Add summary sheet
ws_sum = wb.create_sheet("Summary")
ws_sum["A1"] = "MINSP-005 Test Case Summary"
ws_sum["A1"].font = Font(bold=True, size=14)

total = len(all_cases)
counts = df["Priority"].value_counts()
ws_sum["A3"] = "Priority"; ws_sum["B3"] = "Count"
ws_sum["A3"].font = Font(bold=True); ws_sum["B3"].font = Font(bold=True)
for i, (prio, cnt) in enumerate(counts.items(), start=4):
    ws_sum[f"A{i}"] = prio
    ws_sum[f"B{i}"] = cnt
ws_sum[f"A{i+2}"] = "TOTAL"; ws_sum[f"B{i+2}"] = total
ws_sum[f"A{i+2}"].font = Font(bold=True); ws_sum[f"B{i+2}"].font = Font(bold=True)

ws_sum.column_dimensions["A"].width = 20
ws_sum.column_dimensions["B"].width = 15

wb.save(output_path)
print(f"\nFile Excel saved: {output_path}")
print(f"Total: {total} test cases")
